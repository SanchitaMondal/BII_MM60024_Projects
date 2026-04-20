

import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from tqdm import tqdm



IMAGE_DIR = r"C:\Users\Prawin\Downloads\skindata\test\malignant"   


OUTPUT_FILE = "abcd_features.xlsx"
SUPPORTED   = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"}


#  Segmentation

def get_mask(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (7, 7), 0)
    _, mask = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (15, 15))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, k, iterations=2)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN,  k, iterations=1)
    n, labels, stats, _ = cv2.connectedComponentsWithStats(mask, 8)
    if n > 1:
        biggest = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA])
        mask = np.uint8(labels == biggest) * 255
    return mask


def get_contour(mask):
    cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    return max(cnts, key=cv2.contourArea) if cnts else None


#  A: Asymmetry

def asymmetry_score(mask):
    """0.0 = perfectly symmetric, 1.0 = fully asymmetric"""
    cnt = get_contour(mask)
    if cnt is None:
        return np.nan
    m = cv2.moments(cnt)
    if m["m00"] == 0:
        return 0.0
    cx, cy = int(m["m10"]/m["m00"]), int(m["m01"]/m["m00"])
    theta  = 0.5 * np.arctan2(2*m["mu11"]/m["m00"],
                               (m["mu20"]-m["mu02"])/m["m00"])
    rot    = cv2.getRotationMatrix2D((cx, cy), np.degrees(theta), 1.0)
    h, w   = mask.shape
    aligned = cv2.warpAffine(mask, rot, (w, h))

    scores = []
    for axis in [0, 1]:
        parts = np.array_split(aligned, 2, axis=axis)
        a, b  = parts[0].astype(bool), parts[1].astype(bool)
        b = np.flipud(b) if axis == 0 else np.fliplr(b)
        r, c = min(a.shape[0], b.shape[0]), min(a.shape[1], b.shape[1])
        union = np.logical_or(a[:r, :c], b[:r, :c]).sum()
        inter = np.logical_and(a[:r, :c], b[:r, :c]).sum()
        scores.append(1 - inter/union if union > 0 else 0)
    return round(float(np.mean(scores)), 4)


#  B: Border

def border_irregularity(mask):
    """Compactness: circle=1.0, irregular lesion > 1.0"""
    cnt = get_contour(mask)
    if cnt is None:
        return np.nan
    area = cv2.contourArea(cnt)
    peri = cv2.arcLength(cnt, True)
    if area == 0:
        return np.nan
    return round((peri ** 2) / (4 * np.pi * area), 4)


# C: Color 

def color_variance(img, mask):
    """Mean std-dev of pixel values inside lesion (higher = more color variation)"""
    pixels = img[mask > 0].astype(np.float32)
    if len(pixels) == 0:
        return np.nan
    return round(float(pixels.std(axis=0).mean()), 4)


def color_entropy(img, mask):
    """Shannon entropy of grayscale histogram inside lesion (bits)"""
    gray   = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    pixels = gray[mask > 0]
    if len(pixels) == 0:
        return np.nan
    hist, _ = np.histogram(pixels, bins=32, range=(0, 256), density=True)
    hist    = hist[hist > 0]
    return round(float(-np.sum(hist * np.log2(hist))), 4)


# D: Diameter

def diameter_pixels(mask):
    """Equivalent circular diameter of the lesion in pixels"""
    cnt = get_contour(mask)
    if cnt is None:
        return np.nan
    area = cv2.contourArea(cnt)
    return round(2 * np.sqrt(area / np.pi), 4) if area > 0 else np.nan


def lesion_area_ratio(mask):
    """Fraction of the total image covered by the lesion (0–1)"""
    total = mask.shape[0] * mask.shape[1]
    return round(float((mask > 0).sum() / total), 4) if total > 0 else np.nan


# Per-image extraction 

def extract(path):
    img = cv2.imread(str(path))
    if img is None:
        return None
    mask = get_mask(img)
    return {
        "filename":          path.name,
        "A_asymmetry":       asymmetry_score(mask),
        "B_border_compact":  border_irregularity(mask),
        "C_color_variance":  color_variance(img, mask),
        "C_color_entropy":   color_entropy(img, mask),
        "D_diameter_px":     diameter_pixels(mask),
        "D_area_ratio":      lesion_area_ratio(mask),
    }


# Excel writer

HEADERS = {
    "filename":         "Image File",
    "A_asymmetry":      "A – Asymmetry (0–1)",
    "B_border_compact": "B – Border Compactness",
    "C_color_variance": "C – Color Variance",
    "C_color_entropy":  "C – Color Entropy (bits)",
    "D_diameter_px":    "D – Diameter (px)",
    "D_area_ratio":     "D – Area Ratio (0–1)",
}

COLS = list(HEADERS.keys())


def write_excel(records, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ABCD Features"

    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    alt_fill    = PatternFill("solid", fgColor="EAF1FB")

    for c, key in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=HEADERS[key])
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center

    for r, rec in enumerate(records, 2):
        fill = alt_fill if r % 2 == 0 else PatternFill()
        for c, key in enumerate(COLS, 1):
            cell = ws.cell(row=r, column=c, value=rec.get(key, ""))
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = center
            cell.fill      = fill

    ws.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 22

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"

    wb.save(out_path)
    print(f"\n Saved  {out_path}  ({len(records)} rows)")


# Main 

if __name__ == "__main__":
    folder = Path(IMAGE_DIR)
    if not folder.exists():
        raise FileNotFoundError(f"Directory not found: {folder}\n"
                                f"Please update IMAGE_DIR on line 18.")

    images = sorted(p for p in folder.rglob("*") if p.suffix.lower() in SUPPORTED)
    print(f"Found {len(images)} image(s) in: {folder}")

    records, failed = [], []
    for path in tqdm(images, desc="Extracting", unit="img"):
        try:
            feat = extract(path)
            if feat:
                records.append(feat)
            else:
                failed.append(path.name)
        except Exception as e:
            failed.append(f"{path.name} — {e}")

    if records:
        write_excel(records, OUTPUT_FILE)
    else:
        print("No features extracted. Check your IMAGE_DIR path.")

    if failed:
        print(f"\n  Skipped {len(failed)} file(s): {failed[:5]}")